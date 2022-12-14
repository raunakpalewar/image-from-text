import requests
from bs4 import BeautifulSoup
import openpyxl
import urllib.request
import wikipedia
import random
from PIL import Image
import matplotlib.pyplot as plt
import os
from skimage import io
import cv2


dataframe = openpyxl.load_workbook("F:\\PROJECTS\\project procohat\\pro\\book7000.xlsx")
 
# Define variable to read sheet
dataframe1 = dataframe.active
for row in range(0, dataframe1.max_row):
    for col in dataframe1.iter_cols(1, dataframe1.max_column):
        i=1
        j="1"
        print(col[row].value)
        word=(col[row].value)
        while(col[row].value !=""):
            params = {   
            "q": word,
            "tbm": "isch", 
            "content-type": "image/png",
            "dpi":1200,}

            html = requests.get("https://www.google.com/search?q", params=params)
            

            soup = BeautifulSoup(html.text, 'html.parser')

            img= soup.find("img")
            
            img=img.find_next('img')
            # fig, ax = plt.subplots()
            # fig.savefig('myimage.svg', format='svg', dpi=1200)
            print(img['src'])  
            
            # SAVE_FOLDER = 'imagess'
            # os.mkdir(SAVE_FOLDER)
            # width,height = 'some_w', 'some_h'
            # req = io.imread(img['src'])


            try:
                urllib.request.urlretrieve(img['src'], word+"_img_.png")
            # cv2.imwrite(img['src'], word+"_img_.png", cv2.resize(req, (width,height)))
            except:
                e=open("error.txt","a")
                e.write(word+"\n")

                print("Error\n")

            # dimensions = (500, 500)
            # q = Image.open(word+"_img_.png")
            # q.thumbnail(dimensions)
            # q.show()
            # try:
            #     p = wikipedia.page(word,auto_suggest=False)
            # except wikipedia.exceptions.DisambiguationError as e:
            #     break
            # except wikipedia.exceptions.PageError as f:
            #     break
            # print(p.url)
            # print(p.title)
            # content = p.content # Content of page.
            # x=0
            # text=""
            # for j in content:
            #     if x<2:
            #         text+=j
            #         if(j=='.'):
            #             x+=1
            # with open(word+"_text.txt","w") as f:
            #     f.write(text)

            i=i+1             
            if(i>=1):
                break