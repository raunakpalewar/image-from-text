import os
import json 
import requests 
import openpyxl
import urllib.request
from bs4 import BeautifulSoup # to parse HTML
from skimage import io
import cv2


# user can input a topic and a number
# download first n images from google image search

GOOGLE_IMAGE = 'https://www.google.com/search?q'

# The User-Agent request header contains a characteristic string 
# that allows the network protocol peers to identify the application type, 
# operating system, and software version of the requesting software user agent.
# needed for google search


SAVE_FOLDER = 'images7000_'

def main():
    if not os.path.exists(SAVE_FOLDER):
        os.mkdir(SAVE_FOLDER)
    download_images()

def download_images():
    dataframe = openpyxl.load_workbook("book7000.xlsx")
    a=0
 
    # Define variable to read sheet
    dataframe1 = dataframe.active
    for row in range(0, dataframe1.max_row):
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            i=1
            a=a+1
            j="1"
            data=''
            print(col[row].value)
            data=(col[row].value)
            while(col[row].value !=""):
                usr_agent = {
                'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.11 (KHTML, like Gecko) Chrome/23.0.1271.64 Safari/537.11',
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
                'Accept-Charset': 'ISO-8859-1,utf-8;q=0.7,*;q=0.3',
                'Accept-Encoding': 'none',
                'Accept-Language': 'en-US,en;q=0.8',
                'Connection': 'keep-alive',
                "tbm": "isch", 
                "content-type": "image/png",
                "q":data,
                "dpi":1200,}

                # ask for user input
                n_images = 1
                print('Start searching...')

                # get url query string
                searchurl = GOOGLE_IMAGE + data
                print(searchurl)

                # request url, without usr_agent the permission gets denied
                html = requests.get(searchurl, params=usr_agent)

                # find all divs where class='rg_meta'
                soup = BeautifulSoup(html.text, 'html.parser')
                img= soup.find('img')
            
                img=img.find_next('img')
      
                print(img["src"]) 
                # extract the link from the div tag
                imagelinks=img["src"]
                


                # open image link and save as file
                response = requests.get(imagelinks)
                
                try:

                    imagename = SAVE_FOLDER + '/' + str(a) + '.png'
                    with open(imagename, 'wb') as file:
                        file.write(response.content)
                        print('Done\n\n')

                    print('All Done\n\n')
                except:  
                    e=open("error.txt","a")
                    e.write(str(a)+".  "+data+"\n\n\n\n")

                    print("Error\n\n")

                i=i+1             
                if(i>=1):
                    break
                    


if __name__ == '__main__':
    main()
